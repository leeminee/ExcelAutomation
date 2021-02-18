import datetime
import openpyxl

sample_items = [(datetime.datetime.strptime('2021-02-18', '%Y-%m-%d'), 'Product Name 1', 5, 10000)]

def write_to_file(filepath):
    wb = openpyxl.Workbook()
    for sheet in wb.sheetnames:
        wb.remove(wb[sheet])

        ws = wb.create_sheet(title='2021', index=0)

        ws.merge_cells('A1:F1')
        ws['A1'] = 'Inventory List'
        cell = ws['A1']
