from openpyxl import load_workbook

# data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("C:/test/과일.xlsx", data_only=True)
# 시트 이름으로 불러오기
load_ws = load_wb['Sheet1']

# 셀 주소로 값 출력
print(load_ws['A1'].value)

# 셀 좌표로 값 출력
print(load_ws.cell(1,2).value)

print('\n------지정한 셀 출력------')
get_cells = load_ws['A1':'D2']
for row in get_cells:
    for cell in row:
        print(cell.value)

print('\n------모든 행 단위로 출력------')
for row in load_ws.rows:
    print(row)

print('\n------모든 열 단위로 출력------')
for column in load_ws.columns:
    print(column)

print('\n------모든 행과 열 출력------')
all_values = []
for row in load_ws.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)
print(all_values)