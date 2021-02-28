from openpyxl import Workbook

write_wb = Workbook()

write_ws = write_wb.create_sheet('Alarm 현황')
write_ws['C2'] = 'Host'
write_ws.cell(2,4,'Message1')
write_ws.cell(2,5,'Message2')
write_ws.cell(2,6,'합계')

write_wb.save('C:/test/Alarm현황.xlsx')

