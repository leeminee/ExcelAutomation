from tkinter import *
from tkinter import ttk
from openpyxl import Workbook
from Altibase import altibase_conn
# from Oracle import select_all
import datetime

window = Tk()
window.title("Automation GUI Test")

startDate, endDate, path = StringVar(), StringVar(), StringVar()

now = datetime.datetime.now()
print('now : %s' % now)
# nowDate = now.strftime('%Y-%m-%d %H:%M:%S')
nowDate = now.strftime('%Y%m%d%H%M%S')
print("nowDate : " + nowDate)


def excelDownload():
    write_wb = Workbook()

    a = altibase_conn()

    write_ws = write_wb.active
    write_ws['C2'] = 'Host'
    write_ws.cell(2, 4, 'Usage')
    write_ws.cell(2, 5, 'Total')
    write_ws.cell(2, 6, '합계')


    write_wb.save(path.get() + "/보고서관리_("+nowDate+").xlsx")


ttk.Label(window, text="startDate : ").grid(row=0, column=0, padx=10, pady=10)
ttk.Label(window, text="endDate : ").grid(row=1, column=0, padx=10, pady=10)
ttk.Label(window, text="path to save : ").grid(row=2, column=0, padx=10, pady=10)
ttk.Entry(window, textvariable=startDate).grid(row=0, column=1, padx=10, pady=10)
ttk.Entry(window, textvariable=endDate).grid(row=1, column=1, padx=10, pady=10)
ttk.Entry(window, textvariable=path).grid(row=2, column=1, padx=10, pady=10)
ttk.Button(window, text="Excel Download", command=excelDownload).grid(row=3, column=1, padx=10, pady=10)

window.mainloop()
